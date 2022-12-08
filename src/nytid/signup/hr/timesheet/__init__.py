"""Generate timesheets for TAs"""

import datetime
import io
import pkgutil
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment

def test():
    """Tests the module"""
    name = "Alexander Baltatzis"
    email = "alba@kth.se"
    hourly_salary = 150
    course_leader = ('Daniel Bosk', 'dbosk@kth.se')
    HoD = "Karl Meinke"
    events = []
    events.append({"datum":"2022-12-03",
                  'tid':"8-10",
                  'kurskod':'DD1321',
                  'typ':"handl",
                  'timmar':2,
                  'koeff':1.33,
                  'omr_tid':2*1.33,
                  'belopp':hourly_salary*2*1.33})
    events.append({"datum":"2022-12-04",
                  'tid':"8-10",
                  'kurskod':'DD1321',
                  'typ':"övning",
                  'timmar':2,
                  'koeff':3,
                  'omr_tid':2*3,
                  'belopp':hourly_salary*2*3})
    events.append({"datum":"2022-12-05",
                  'tid':"8-10",
                  'kurskod':'DD1321',
                  'typ':"handl",
                  'timmar':2.1,
                  'koeff':1.33,
                  'omr_tid':2.1*1.33,
                  'belopp':hourly_salary*2.1*1.33})
    events.append({"datum":"2022-12-05",
                  'tid':"10-12",
                  'kurskod':'DD1310',
                  'typ':"handl",
                  'timmar':2,
                  'koeff':1.33,
                  'omr_tid':2*1.33,
                  'belopp':hourly_salary*2*1.33})

    make_excel(name,      
               email,     
               events,
               course_leader,
               HoD,
               course_leader_signature="signature.png"
               )


def make_excel(name, email, events,
               course_leader, HoD,
               org = "JH", project = "1102",
               hourly_salary = 150,
               output = None,
               course_leader_signature = None,
               HoD_signature = None,
               logo = "kth.png"):
    """
    Generates a time report for a TA:
    - `name` and `email` are name and email for the TA.
    - `events` is a list of dictionaries, each dictionary containing:
        - datum: str
        - tid: str
        - kurskod: str
        - typ: str
        - timmar: float, koeff: float, omr_tid: float
        - belopp: float
    - `course_leader` is a tuple (name: str, email: str)
    - `HoD` the name of the Head of Department
    - `org`, `project` is the organization and project, both strings.
    - `hourly_salary` is the hourly salary, float.
    - `output` is the desired output (xlsx) filename.
    - `course_leader_signature` is an image (or path to one) containing the 
      course responsible's signature.
    - `HoD_signature` is an image (or path to one) containing the Head of 
      Department's signature.
    - `logo` is an image (or path to one) containing the logotype of the 
      university.
    """
    login = email.replace("@kth.se", "")
    if not output:
        output = login + "_tid_" + datetime.date.today().strftime("%Y-%m-%d.xlsx")

    wb = Workbook()
    ark = wb.active


    #############################################################
    # Logo
    ark.title = login + " " + datetime.date.today().strftime("%Y-%b")
    try:
        logo = Image(logo)
        scale = 80/logo.height
        logo.height *= scale
        logo.width *= scale
        ark.add_image(logo, "A1")
    except:
        pass

    #############################################################
    # kolumnstorlekar
    ark.column_dimensions['A'].width = 16  # 'Schemalagd tid'
    ark.column_dimensions['B'].width = 9   # 'Typ'           
    ark.column_dimensions['C'].width = 7   # 'timmar'      
    ark.column_dimensions['D'].width = 8   # 'koeff'         
    ark.column_dimensions['E'].width = 15  # 'omräknad tid'        
    ark.column_dimensions['F'].width = 7   # 'Timlön'     
    ark.column_dimensions['G'].width = 9   # 'Belopp'        
    ark.column_dimensions['H'].width = 9

    #############################################################
    # Börja på rad 6
    rad = "6"
    ark['A' + rad] = "Timredovisning"
    ark['D' + rad] = "Namn"
    ark['E' + rad] = name
    ark['E' + rad].fill = PatternFill(start_color="00EEECE1", end_color="00EEECE1", fill_type="solid")

    rad = incr(rad)
    ark['D' + rad] = 'epost'
    ark['E' + rad] = email
    ark['E' + rad].fill = PatternFill(start_color="00EEECE1", end_color="00EEECE1", fill_type="solid")

    rad = incr(rad, 2)
    ark['A' + rad] = 'Kurskod'
    ark['B' + rad] = ''
    kurskoder = []
    for kol in events:
        if 'kurskod' in kol and kol['kurskod'] not in kurskoder:
            ark['B' + rad].value += kol['kurskod'] + " "
            kurskoder.append( kol['kurskod'] )
    
    rad = incr(rad)
    ark['A' + rad] = 'Timmar ska anges inklusive förberedelsetid enligt schablon'
    rad = incr(rad)
    ark['A' + rad] = 'Ange typ av undervisning övning, handledning'

    rad = incr(rad, 2)
    ark['A' + rad] = 'Schemalagd tid'
    ark['B' + rad] = 'Typ'           
    ark['C' + rad] = 'Timmar'        
    ark['D' + rad] = 'koeff'         
    ark['E' + rad] = 'Omräknad tid'
    ark['F' + rad] = 'Timlön'        
    ark['G' + rad] = 'Belopp'
    
    for kol in ['C', 'D', 'E', 'F', 'G']:
        ark[kol+rad].alignment = Alignment(horizontal="right")
    

    #############################################################
    # Summering på sista raden 
    rad = incr(rad)   
    sist = incr(rad, len(events))
    ark['E'+sist].font = Font(bold=True)  
    ark['G'+sist].font = Font(bold=True)  
    ark['E'+sist].fill = PatternFill(start_color="00EEECE1", end_color="00EEECE1", fill_type="solid")
    #ark['G'+sist].fill = PatternFill(start_color="00C0C0C0", end_color="00C0C0C0", fill_type="solid")

    #############################################################
    # Matris med timredovisningen
    for i, kol in enumerate(events):
        ark['A'+rad].value = f"{kol['datum']} {kol['tid']:>5}"
        ark['B'+rad] = kol['typ']      
        ark['C'+rad] = kol['timmar']
        ark['D'+rad] = kol['koeff']    
        ark['E'+rad].value = round(kol['omr_tid'], 1)
        ark['F'+rad].value = hourly_salary
        ark['G'+rad].value = round(kol['belopp'], 1)

        if i % 2 == 0:
            for kol in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                ark[kol+rad].fill = PatternFill(start_color="00E0E0E0", end_color="00E0E0E0", fill_type="solid")
        
        if i == 0:
            tidsumma = "=ROUNDUP(SUM(E"+rad 
            ark['G'+sist].value = '=G'+rad 
        else:
            tidsumma += ',E'+rad 
            ark['G'+sist].value += '+G'+rad

        rad = incr(rad)

    ark['E'+sist].value = tidsumma + '),1)'

            
    #############################################################
    # Kontering
    rad = incr(sist, 3)
    ark['A'+rad].value = "Kontering"
    rad = incr(rad)
    ark['A'+rad].value = "Org.enhet"
    ark['A'+rad].fill  = PatternFill(start_color="00EEECE1", end_color="00EEECE1", fill_type="solid")
    ark['B'+rad].value = "Projekt"
    ark['B'+rad].fill = PatternFill(start_color="00EEECE1", end_color="00EEECE1", fill_type="solid")
    rad = incr(rad)
    ark['A'+rad].value = org
    ark['A'+rad].fill  = PatternFill(start_color="00EEECE1", end_color="00EEECE1", fill_type="solid")
    ark['B'+rad].value = project
    ark['B'+rad].fill  = PatternFill(start_color="00EEECE1", end_color="00EEECE1", fill_type="solid")

    #############################################################
    # Underskrift
    rad = incr(rad, 4)

    ark['A'+rad].value = "_______________________________________"
    ark['E'+rad].value = "_______________________________________"

    if HoD_signature:
        HoD_signature = Image(HoD_signature)
        scale = 60/HoD_signature.height
        HoD_signature.width *= scale
        HoD_signature.height *= scale
        ark.add_image(HoD_signature, "A"+incr(rad, -2))

    if course_leader_signature:
        course_leader_signature = Image(course_leader_signature)
        scale = 60/course_leader_signature.height
        course_leader_signature.width *= scale
        course_leader_signature.height *= scale
        ark.add_image(course_leader_signature, "E"+incr(rad, -2))

    rad = incr(rad)

    ark['A'+rad].value = "Ekonomisk attest " + HoD
    ark['E'+rad].value = "Kursansvarig"

    rad = incr(rad)
    ark['E'+rad].value = course_leader[0]
    rad = incr(rad)
    ark['E'+rad].value = course_leader[1]

    rad = incr(rad, 2)
    ark['A'+rad].value = "Underskriven blankett lämnas till HR"

    wb.save(output)

def incr(rad, i=1):
    """
    Bumps the row of a cell. Takes a string containing a number,
    increase by `i`.
    """
    return str(int(rad)+i)


#################################################################
#
# main - test
#
if __name__ == "__main__":
    test()
